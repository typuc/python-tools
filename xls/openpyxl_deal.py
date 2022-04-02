import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill
wb = openpyxl.load_workbook("/Users/leo/Desktop/th.xlsx")

# https://openpyxl.readthedocs.io/en/stable/tutorial.html#
# https://zhuanlan.zhihu.com/p/259583430
my_sheet = wb['test']
sheet = wb.create_sheet(index=0, title="Report")
# cell1 = my_sheet['A2']
# cell2 = my_sheet['B2']
# print(cell1.value, cell2.value)
# print(my_sheet.cell(row=1, column=4).value)
# print(my_sheet.max_row)
# print(my_sheet.max_column)
orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")

r = 0
c = 0
shop_list = list()
for row in my_sheet.rows:
    r = r + 1
    cn = column_index_from_string('I')
    name = my_sheet.cell(row=r, column=cn).value
    if name == '李蕾':
        shop_id = my_sheet.cell(row=r, column=2).value
        sale_date = my_sheet.cell(row=r, column=3).value
        #填充颜色
        my_sheet.cell(row=r, column=cn).fill = orange_fill
        # print(shop_id, name)
        shop_list.append(shop_id)
print(list(set(shop_list)))
print(len(list(set(shop_list))))
wb.save("/Users/leo/Desktop/th.xlsx")
wb.close()
